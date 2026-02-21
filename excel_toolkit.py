import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def clean_excel(input_file, output_file):
    print("📊 Loading file...")
    
    # Load Excel
    df = pd.read_excel(input_file)

    # Standardize column names
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

    # Remove duplicate rows
    df = df.drop_duplicates()

    # Remove fully empty rows
    df = df.dropna(how="all")

    # Save cleaned data
    df.to_excel(output_file, index=False, sheet_name="Cleaned_Data")

    print("✅ Data cleaned successfully!")

    # Add summary sheet
    add_summary(output_file, df)


def add_summary(file_path, df):
    print("📈 Generating summary sheet...")

    total_rows = len(df)
    total_columns = len(df.columns)
    missing_values = df.isnull().sum().sum()

    summary_data = {
        "Metric": ["Total Rows", "Total Columns", "Total Missing Values"],
        "Value": [total_rows, total_columns, missing_values]
    }

    summary_df = pd.DataFrame(summary_data)

    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # Style header bold
    wb = load_workbook(file_path)
    ws = wb["Summary"]

    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save(file_path)

    print("📊 Summary sheet added!")


if __name__ == "__main__":
    input_file = "input.xlsx"
    output_file = "cleaned_output.xlsx"

    clean_excel(input_file, output_file)
